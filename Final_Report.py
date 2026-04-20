import csv
import os
import zipfile
from typing import List, Tuple
from xml.sax.saxutils import escape

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except Exception:  # pragma: no cover - runtime fallback
    Workbook = None


def script_dir() -> str:
    return os.path.dirname(os.path.abspath(__file__))


def read_csv_rows(path: str) -> List[List[str]]:
    with open(path, newline="", encoding="utf-8") as f:
        reader = csv.reader(f)
        return [[c if c is not None else "" for c in row] for row in reader]


def sanitize_sheet_name(name: str) -> str:
    bad = ['\\', '/', '*', '?', ':', '[', ']']
    out = "".join("_" if ch in bad else ch for ch in name)
    out = out.strip() or "Sheet"
    return out[:31]


def col_name(index_1_based: int) -> str:
    n = index_1_based
    out = []
    while n > 0:
        n, r = divmod(n - 1, 26)
        out.append(chr(65 + r))
    return "".join(reversed(out))


def is_number(value: str) -> bool:
    if value is None:
        return False
    s = str(value).strip()
    if not s:
        return False
    try:
        float(s)
        return True
    except Exception:
        return False


def worksheet_xml(rows: List[List[str]]) -> str:
    lines = [
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>',
        '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">',
        "<sheetData>",
    ]
    for r_idx, row in enumerate(rows, start=1):
        lines.append(f'<row r="{r_idx}">')
        for c_idx, val in enumerate(row, start=1):
            ref = f"{col_name(c_idx)}{r_idx}"
            txt = "" if val is None else str(val)
            if is_number(txt):
                lines.append(f'<c r="{ref}"><v>{escape(txt)}</v></c>')
            else:
                lines.append(f'<c r="{ref}" t="inlineStr"><is><t>{escape(txt)}</t></is></c>')
        lines.append("</row>")
    lines.append("</sheetData>")
    lines.append("</worksheet>")
    return "".join(lines)


def workbook_xml(sheet_names: List[str]) -> str:
    sheets = []
    for i, n in enumerate(sheet_names, start=1):
        sheets.append(
            f'<sheet name="{escape(n)}" sheetId="{i}" r:id="rId{i}"/>'
        )
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
        'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
        f"<sheets>{''.join(sheets)}</sheets>"
        "</workbook>"
    )


def workbook_rels_xml(sheet_count: int) -> str:
    rels = []
    for i in range(1, sheet_count + 1):
        rels.append(
            f'<Relationship Id="rId{i}" '
            'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" '
            f'Target="worksheets/sheet{i}.xml"/>'
        )
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        f"{''.join(rels)}"
        "</Relationships>"
    )


def root_rels_xml() -> str:
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" '
        'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" '
        'Target="xl/workbook.xml"/>'
        "</Relationships>"
    )


def content_types_xml(sheet_count: int) -> str:
    overrides = [
        '<Override PartName="/xl/workbook.xml" '
        'ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>'
    ]
    for i in range(1, sheet_count + 1):
        overrides.append(
            f'<Override PartName="/xl/worksheets/sheet{i}.xml" '
            'ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>'
        )
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
        '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
        '<Default Extension="xml" ContentType="application/xml"/>'
        f"{''.join(overrides)}"
        "</Types>"
    )


def build_workbook(sheets: List[Tuple[str, List[List[str]]]], out_path: str) -> None:
    if Workbook is not None:
        _build_workbook_openpyxl(sheets, out_path)
        return

    # Fallback path when openpyxl is unavailable.
    sheet_names = [sanitize_sheet_name(n) for n, _ in sheets]
    with zipfile.ZipFile(out_path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("[Content_Types].xml", content_types_xml(len(sheets)))
        zf.writestr("_rels/.rels", root_rels_xml())
        zf.writestr("xl/workbook.xml", workbook_xml(sheet_names))
        zf.writestr("xl/_rels/workbook.xml.rels", workbook_rels_xml(len(sheets)))
        for i, (_, rows) in enumerate(sheets, start=1):
            zf.writestr(f"xl/worksheets/sheet{i}.xml", worksheet_xml(rows))


def _is_numeric_header(header: str) -> bool:
    h = (header or "").strip().lower()
    return any(
        k in h
        for k in [
            "qty",
            "quantity",
            "price",
            "cost",
            "rate",
            "amount",
            "volume",
            "area",
            "length",
            "width",
            "wastage",
            "%",
            "id",
            "express_id",
        ]
    )


def _is_currency_header(header: str) -> bool:
    h = (header or "").strip().lower()
    # Apply accounting style only for final rate table money columns.
    return h in {"price (₹)", "total amount (₹)"}


def _is_percent_header(header: str) -> bool:
    h = (header or "").strip().lower()
    return "wastage" in h or "%" in h


def _to_number(value):
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None
    # Accept already-formatted Indian currency/number strings.
    neg = False
    if s.startswith("(") and s.endswith(")"):
        neg = True
        s = s[1:-1]
    s = (
        s.replace("₹", "")
        .replace(",", "")
        .replace(" ", "")
        .replace("\u00a0", "")
    )
    if s in ("", "-", "--"):
        return None
    try:
        if "." in s:
            v = float(s)
        else:
            v = int(s)
        return -v if neg else v
    except Exception:
        return None


def _build_workbook_openpyxl(sheets: List[Tuple[str, List[List[str]]]], out_path: str) -> None:
    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    thin = Side(style="thin", color="D9DEE5")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill(fill_type="solid", fgColor="E9EEF6")
    section_fill = PatternFill(fill_type="solid", fgColor="F7FAFF")
    total_fill = PatternFill(fill_type="solid", fgColor="EAF7EE")

    header_font = Font(name="Calibri", size=11, bold=True, color="1F2937")
    body_font = Font(name="Calibri", size=10, color="111827")
    total_font = Font(name="Calibri", size=10, bold=True, color="0F5132")

    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_right = Alignment(horizontal="right", vertical="center", wrap_text=True)

    accounting_inr_fmt = '_-₹* #,##,##0.00_-;[Red]-₹* #,##,##0.00_-;_-₹* "-"??_-;_-@_-'
    num_3dec_fmt = "0.000"
    num_2dec_fmt = "0.00"

    for sheet_name, rows in sheets:
        ws = wb.create_sheet(title=sanitize_sheet_name(sheet_name))

        if not rows:
            continue

        max_cols = max((len(r) for r in rows), default=0)
        headers = [str(h or "") for h in (rows[0] if rows else [])]

        col_widths = [len(h) for h in headers] + [10] * max(0, max_cols - len(headers))

        for r_idx, row in enumerate(rows, start=1):
            row_values = list(row) + [""] * (max_cols - len(row))
            for c_idx, value in enumerate(row_values, start=1):
                hdr = headers[c_idx - 1] if c_idx - 1 < len(headers) else ""
                is_num_col = _is_numeric_header(hdr)

                num_value = _to_number(value)
                cell_value = num_value if (num_value is not None and is_num_col) else ("" if value is None else str(value))

                cell = ws.cell(row=r_idx, column=c_idx, value=cell_value)
                cell.border = border
                cell.font = body_font

                if r_idx == 1:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = align_center
                else:
                    cell.alignment = align_right if is_num_col else align_left

                    code_val = str(row_values[0] if row_values else "").strip()
                    material_val = str(row_values[3] if len(row_values) > 3 and row_values[3] is not None else "").strip()

                    if code_val.startswith("Total for "):
                        cell.fill = total_fill
                        cell.font = total_font
                    elif material_val == "" and "Summary" in ws.title:
                        cell.fill = section_fill

                if isinstance(cell_value, (int, float)) and r_idx > 1:
                    if is_num_col:
                        hdr_l = hdr.lower()
                        if "id" in hdr_l and hdr_l == "express_id":
                            cell.number_format = "0"
                        elif _is_percent_header(hdr):
                            cell.number_format = num_2dec_fmt
                        elif _is_currency_header(hdr):
                            cell.number_format = accounting_inr_fmt
                        else:
                            cell.number_format = num_3dec_fmt

                text_len = len("" if cell_value is None else str(cell_value))
                if c_idx - 1 < len(col_widths):
                    col_widths[c_idx - 1] = max(col_widths[c_idx - 1], text_len)

            # Add vertical breathing room so wrapped text is fully readable.
            if r_idx == 1:
                ws.row_dimensions[r_idx].height = 24
            else:
                ws.row_dimensions[r_idx].height = 22

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        for c_idx in range(1, max_cols + 1):
            letter = get_column_letter(c_idx)
            header = headers[c_idx - 1].lower() if c_idx - 1 < len(headers) else ""
            if "description" in header or "item" in header or "material" in header:
                width_cap = 72
            elif "family" in header or "type" in header:
                width_cap = 42
            else:
                width_cap = 28

            width = min(max(12, col_widths[c_idx - 1] + 3), width_cap)
            ws.column_dimensions[letter].width = width

    wb.save(out_path)


def ensure_inputs(base: str, required_csvs: List[str]) -> None:
    # Regenerate only if one or more source CSVs are missing.
    missing = [p for p in required_csvs if not os.path.exists(p)]
    if not missing:
        return

    wall_py = os.path.join(base, "Wall.py")
    rcc_py = os.path.join(base, "RCC.py")
    if os.path.exists(wall_py):
        os.system(f'python "{wall_py}"')
    if os.path.exists(rcc_py):
        os.system(f'python "{rcc_py}"')


def main() -> None:
    base = script_dir()
    wall_sheet = os.path.join(base, "Wall_Sheet.csv")
    wall_summary = os.path.join(base, "Wall_Sheet_Summary.csv")
    rcc_sheet = os.path.join(base, "RCC_sheet.csv")
    rcc_summary = os.path.join(base, "RCC_Summary.csv")

    required = [wall_sheet, wall_summary, rcc_sheet, rcc_summary]
    ensure_inputs(base, required)
    missing = [p for p in required if not os.path.exists(p)]
    if missing:
        raise SystemExit(f"Missing input CSV(s): {', '.join(missing)}")

    sheets = [
        ("Wall_Sheet", read_csv_rows(wall_sheet)),
        ("Wall_Summary", read_csv_rows(wall_summary)),
        ("RCC_Sheet", read_csv_rows(rcc_sheet)),
        ("RCC_Summary", read_csv_rows(rcc_summary)),
    ]

    out_xlsx = os.path.join(base, "final.xlsx")
    build_workbook(sheets, out_xlsx)
    print(f"Created {out_xlsx} with 4 sheets.")
    print("Note: CSV format cannot contain multiple sheets, so output is .xlsx.")


if __name__ == "__main__":
    main()
